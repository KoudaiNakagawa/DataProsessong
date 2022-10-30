import openpyxl     #openpyxl というモジュールを使う宣言をするよ。

excel_path = "./13000-15.0b/13_2021.xlsx"                   #Excelのファイルのパスを覚えておくよ。
sheetName = "13_2021"                                       #オリジナルデータのシート名を覚えておくよ。
wb = openpyxl.load_workbook(excel_path, data_only = True)   #Excelのファイルとオプションを覚えておくよ。データは関数で処理した後の数値だよ。
inputSheet = wb[sheetName]                                  #入力するシート名を覚えておくよ。

objectClassCell = 4         #分類する要素の列を覚えておくよ。
cellRange = range(5)        #いくつデータを取り出すか覚えておくよ。cellRangeには、[0, 1, 2, 3, 4]というリストが入っているよ。
y = 1                       #行の開始位置を覚えておくよ。

objectList = []             #空のリストを作るよ。
for x in cellRange:                                                                     #cellRangeから、順番に数字を取り出して x に代入するループを回すよ。cellRangeには、[0, 1, 2, 3, 4]というリストが入っているよ。
    objectList.append(str(inputSheet.cell(row=y, column= x + objectClassCell).value))   #取り出した x と、 y でCellの位置を指定して、値を取ってくるよ。値は、objectList に順番に入れるよ。

y+=1                        #次の行に移るために、y に +1 するよ。   
objectClass = ""            #分類する要素は、最初は空の文字列だよ。
offset = 0                  #書き込みを2行目から始めるための数を覚えておくよ。

while True:                 #無限ループ[1]を回すよ。止めるときは別の命令で止めてね。
    objectClass = str(inputSheet.cell(row= y, column= objectClassCell).value)      #次の要素名を objectClass に代入するよ。
    
    if objectClass == "None":                  #もし読み込む要素が無くなったら(objectClassがNoneだったら)、
        break                                  #ループ[1]を終了するよ。

    try:                                       #エラー覚悟で下の命令にチャレンジしてね。       
        outputsheet = wb[objectClass]               #書き込むシートは、作ったシートだよ。
    except:                                    #エラーが出た(書き込むシートが無かった)ら、
        wb.create_sheet(objectClass)                #objectClassと同じ名前のシートを作るよ。
        outputsheet = wb[objectClass]               #書き込むシートは、作ったシートだよ。

    for x in cellRange:                                                 #cellRangeから、順番に数字を取り出して x に代入するループを回すよ。cellRangeには、[0, 1, 2, 3, 4]というリストが入っているよ。
        outputsheet.cell(row= 1, column= x+1, value= objectList[x])     #objectList から見出しを書き込むよ。

    while True:             #無限ループ[2]を回すよ。止めるときは別の命令で止めてね。

        if str(inputSheet.cell(row= y, column= objectClassCell).value) != objectClass:   #もし分類する要素と持ってきた要素が違ったら、
            offset = y-2                                                                 #書き込みを2行目から始めるための数を足して、
            break                                                                        #[2]のループを終了するよ。

        for x in cellRange:                                                                 #cellRangeから、順番に数字を取り出して x に代入するよ。cellRangeには、[0, 1, 2, 3, 4]というリストが入っているよ。
            inputCell = inputSheet.cell(row= y, column= x + objectClassCell).value          #取り出した x と、 y でCellの位置を指定するよ。値は、inputCellに代入するよ。
            outputsheet.cell(row= y-offset, column= x+1, value= inputCell)                         #inputCellの値を、 要素の名前を持つシートに書き込むよ。            
        y+=1                #次の行に移るために、y に +1 するよ。 

wb.save(excel_path)         #指定されたExcelのファイルを保存するよ。
wb.close()                  #指定されたExcelのファイルを閉じるよ。