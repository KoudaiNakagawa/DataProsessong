import folium
import openpyxl

city = "町田市"



coordinateDataPath = "./13000-15.0b/13_2021.xlsx"
coordinateWB = openpyxl.load_workbook(coordinateDataPath, data_only=True)               #座標データのExcelを開くよ。
coordinateInputSheet = coordinateWB[city]                                               #シートは、cityと同じ名前のものだよ。

coordinateDict = {}
sumLat = 0
sumLong = 0
offsetY = 2
y = offsetY
while True:
    coordinateList = []
    if str(coordinateInputSheet.cell(row= y, column= 1).value) == "None":               #もし空のセルを見たら、
        avgLat = sumLat / (y - offsetY)                                                 #緯度の平均を計算して、
        avgLong = sumLong / (y - offsetY)                                               #軽度の平均を計算して、
        break                                                                           #ループを終了してね。

    for x in range(1,6):
        coordinateList.append(coordinateInputSheet.cell(row= y, column= x).value)       #ExcelからX方向にデータを取ってきてリストに格納するよ。
        
    coordinateDict[coordinateList[0]+coordinateList[2]] = [coordinateList[3],coordinateList[4],coordinateList[1]]   #警視庁のデータから参照しやすくするために、Keyに市と丁目、valueに緯度、経度、市区町村コードの形式で辞書に覚えておくよ。
    sumLat += coordinateList[3]                                                                                     #緯度の合計を覚えておくよ。後で平均を出すよ。
    sumLong += coordinateList[4]                                                                                    #軽度の合計を覚えておくよ。後で平均を出すよ。
    y+=1
coordinateWB.close()                                                                    #座標データのExcelを閉じるよ。



mainDataPath = "./R4.xlsx"                                                              
mainWB = openpyxl.load_workbook(mainDataPath, data_only=True)                           #犯罪データのExcelを開くよ。
mainInputSheet = mainWB[city]                                                           #シートは、cityと同じ名前のものだよ。
DataSet = [2, 3, 6, 12, 21, 33, 39]                                                     #それぞれの要素の親の位置を覚えておくよ。

crimeY = 7 
crimeClassList = []
for x1 in DataSet:                                                                      
    crimeList = []
    if x1 == 39:
        break
    elif x1 == 2:
        crimeMasterClass = mainInputSheet.cell(row= crimeY -2 , column= x1).value       
    else:        
        crimeMasterClass = mainInputSheet.cell(row= crimeY -1 , column= x1).value
        for x2 in range(x1+1,DataSet[DataSet.index(x1)+1]):
            crimeList.append(mainInputSheet.cell(row= crimeY, column= x2).value)

    crimeClassList.append([crimeMasterClass, crimeList])                                #要素の親と子を覚えておくよ。



transPattern = str.maketrans({'１':'一', '２':'二', '３':'三', '４':'四', '５':'五', '６':'六', '７':'七', '８':'八', '９':'九'})   #全角アラビア数字から、漢数字に変換するためのパターンを覚えておくよ。


for n in range(len(DataSet)-1):
    cityMap = folium.Map(location=[avgLat,avgLong], zoom_start= 12.5)

    y=8
    crimeClass = crimeClassList[n][0]

    if crimeClass == "None":
        break

    if n == 0:                                                                              #もし総合計のデータをを見ようとしているなら、各要素の合計のデータを取ってきてね。
        while True:                                                 #ループを回すよ。
            cityCrimeList = []
            keyCell = mainInputSheet.cell(row= y, column= 1).value  #セルからデータを取ってきて、
            if "計" in keyCell:                                     #もし「計」という文字が入っていたら、
                break                                               #ループを終了するよ。

            keyName = keyCell.translate(transPattern)               #アラビア漢数字変換。
            
            try:                                                    #エラーが出るかもしれないけどチャレンジしてね。
                coordinateDataList = coordinateDict[keyName]        #辞書から、キーに対応する要素を持ってきてね。
                lat = coordinateDataList[0]
                long = coordinateDataList[1]

                
                text = ""
                for i in range(len(DataSet)-1):
                    cityCrimeList.append(mainInputSheet.cell(row= y, column= DataSet[i]).value) #リストに格納して、
                    text += str(crimeClassList[i][0]) + ":" + str(cityCrimeList[i]) + "件\n"    #テキストを作ってね。
                

                folium.CircleMarker(
                location=[lat, long],
                radius=cityCrimeList[0],
                color='#ff0000',
                fill_color='#0000ff'
                ).add_to(cityMap)                                   #犯罪件数に比例した大きさの円を描いてね。

                folium.Marker(
                location=[lat, long],
                popup= keyName + "\n" + text + ("　"*16)
                ).add_to(cityMap)                                   #テキストを配置してね。


            except KeyError:                                        #エラーが出た要素をコンソールに出力してね。
                print("NotFound :", keyName)

            y+=1

        cityMap.save(city + crimeClass +".html")



    else:                                                                                   #もし種類別のデータを見ようとしているなら、それぞれの子のデータを取ってきてね。
        while True:
            cityCrimeList = []
            keyCell = mainInputSheet.cell(row= y, column= 1).value
            if "計" in keyCell:
                break
            keyName = keyCell.translate(transPattern)
            
            try:
                coordinateDataList = coordinateDict[keyName]
                lat = coordinateDataList[0]
                long = coordinateDataList[1]

                
                text = ""
                for i in range(DataSet[n+1] - DataSet[n] -1):
                    rng = range(DataSet[n], DataSet[n+1])
                    cityCrimeList.append(mainInputSheet.cell(row= y, column= rng[i]).value)
                    text += str(crimeClassList[n][1][i]) + ":" + str(cityCrimeList[i]) + "件\n"


                folium.CircleMarker(
                location=[lat, long],
                radius=cityCrimeList[0],
                color='#ff0000',
                fill_color='#0000ff'
                ).add_to(cityMap)

                folium.Marker(
                location=[lat, long],
                popup= keyName + "\n" + text + ("　"*16)
                ).add_to(cityMap)


            except KeyError:
                print("NotFound :", keyName)

            y+=1

        cityMap.save(city + crimeClass +".html")

print("End")